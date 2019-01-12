import openpyxl

theFile = openpyxl.load_workbook('Customers1.xlsx')
allSheetNames = theFile.sheetnames

print("All sheet names {} " .format(theFile.sheetnames))


def find_specific_cell():
    for row in range(1, currentSheet.max_row + 1):
        for column in "ABCDEFGHIJKL":  # Here you can add or reduce the columns
            cell_name = "{}{}".format(column, row)
            if currentSheet[cell_name].value == "telephone":
                #print("{1} cell is located on {0}" .format(cell_name, currentSheet[cell_name].value))
                print("cell position {} has value {}".format(cell_name, currentSheet[cell_name].value))
                return cell_name

def get_column_letter(specificCellLetter):
    letter = specificCellLetter[0:-1]
    print(letter)
    return letter

#not needed as get_all_values_by_cell_list uses to many resources.
def get_all_cells_from_specific_letter(letter):
    specificCells = []
    for x in range(2, currentSheet.max_row + 1): # 2 is because there is column name on 1
        specificCells.append(letter + str(x))
    return specificCells


def get_all_values_by_cell_letter(letter):
    for row in range(1, currentSheet.max_row + 1):
        for column in letter:
            cell_name = "{}{}".format(column, row)
            #print(cell_name)
            print("cell position {} has value {}".format(cell_name, currentSheet[cell_name].value))



for sheet in allSheetNames:
    print("Current sheet name is {}" .format(sheet))
    currentSheet = theFile[sheet]
    specificCellLetter = (find_specific_cell())
    letter = get_column_letter(specificCellLetter)

    #specificCells = get_all_cells_from_specific_letter(letter)
    get_all_values_by_cell_letter(letter)
   # get_all_values_by_cell_list(specificCells)

#not needed as it uses unnececary resources to find values. get_all_cells_from_specific_letter is more efficient.
def get_all_values_by_cell_list(specificCells):
    for row in range(1, currentSheet.max_row + 1):
        for column in "ABCDEF":  # Here you can add or reduce the columns
            cell_name = "{}{}".format(column, row)
            if cell_name in specificCells:
                print("cell position {} has value {}".format(cell_name, currentSheet[cell_name].value))
#not needed.
def read_all():
    for row in range(1, currentSheet.max_row + 1):
        print(row)
        #for column in range(1, numberOfColumns):
         #   print (column)
        for column in "ABCDEFGHIJKL":  # Here you can add or reduce the columns
            cell_name = "{}{}".format(column, row)
            #print(cell_name)
            print("cell position {} has value {}".format(cell_name, currentSheet[cell_name].value))


